import streamlit as st
import openpyxl

# Streamlitのタイトル
st.title('Excelファイル処理アプリ')

# ファイルアップロード
uploaded_file_icare = st.file_uploader("アイケア変バリスケジュール管理.xlsx ファイルをアップロード", type=['xlsx'])
uploaded_file_ueno = st.file_uploader("上野点眼洗眼帳票.xlsx ファイルをアップロード", type=['xlsx'])

# ボタンが押されたら処理開始
if st.button('項目確認'):
    if uploaded_file_icare is not None and uploaded_file_ueno is not None:
        # ワークブックをロードして、'生産略称'を探す
        icare_workbook = openpyxl.load_workbook(uploaded_file_icare, data_only=True)
        ueno_workbook = openpyxl.load_workbook(uploaded_file_ueno, data_only=True)

        icare_sheet = icare_workbook.active
        ueno_sheet = ueno_workbook.active

        # '生産略称'をシートから探す
        prod_abbrev_cell = None
        for row in icare_sheet.iter_rows():
            for cell in row:
                if cell.value == '生産略称':
                    prod_abbrev_cell = (cell.row, cell.column)
                    break
            if prod_abbrev_cell:
                break

        # '生産略称'の下にある項目を抽出
        prod_abbrev_items = []
        if prod_abbrev_cell:
            for row in icare_sheet.iter_rows(min_row=prod_abbrev_cell[0]+1,
                                            max_col=prod_abbrev_cell[1],
                                            max_row=icare_sheet.max_row):
                for cell in row:
                    if cell.value is not None:
                        prod_abbrev_items.append(cell.value)

        # 見つかった項目と対応するA列の日付を保存する辞書
        found_items_with_A_col_values_based_on_3_chars = {}

        # 40から45列目で各項目を確認
        for row in ueno_sheet.iter_rows(min_row=1, max_row=ueno_sheet.max_row):
            for col in range(40, 46):
                cell_value = row[col-1].value
                if cell_value and any(cell_value.startswith(item[:3]) for item in prod_abbrev_items):
                    found_items_with_A_col_values_based_on_3_chars[cell_value] = row[0].value

        # 結果を表示
        if found_items_with_A_col_values_based_on_3_chars:
            st.write(found_items_with_A_col_values_based_on_3_chars)
        else:
            st.error("一致する項目は見つかりませんでした。")
    else:
        st.error("両方のファイルをアップロードしてください。")
